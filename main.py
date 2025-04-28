import asyncio
import sys

sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')

import json
import os
from dotenv import load_dotenv
import pandas as pd 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from source_url import urls

from crawl4ai import AsyncWebCrawler
from crawl4ai.async_configs import BrowserConfig, CrawlerRunConfig, CacheMode


import google.generativeai as genai
import re
import time

load_dotenv()

async def process_gemini_response(response_text):
    pattern = r'```(?:json)?\s*(.*?)```'
    match = re.search(pattern, response_text, re.DOTALL)
    
    if match:
        json_text = match.group(1).strip()
        try:
            json_data = json.loads(json_text)
            return json_data
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            return None
    else:
        try:
            json_data = json.loads(response_text.strip())
            return json_data
        except json.JSONDecodeError:
            print("Could not extract valid JSON from response")
            return None

async def get_job_info(data, url):
    genai.configure(api_key=os.getenv("GEMINI_API_KEY"))
    model = genai.GenerativeModel("gemini-2.0-flash")
        
    prompt = f"""{url}\n{data}\nHãy phân tích và trích xuất thông tin Việc Làm được giới thiệu ở trang web này, chuyển về dạng JSON với cấu trúc:

    {{
        "jobname - Tên công việc": "$Tên Việc Làm đang được giới thiệu ở trang web này",
        "companyname - Tên Công ty": "$Tên Công ty đang được giới thiệu ở trang web này",
        "url - Địa chỉ web": "$URL href (của công việc đó)",
        "expired - Thời gian hết hạn": "$Thời gian hiệu lực của công việc đang tuyển dụng đó",
        "contact - Thông tin liên hệ": "$Thông tin liên hệ của công ty tuyển dụng (nếu trang web không ghi rõ, hãy đoán hoặc tìm từ website công ty nếu có thể)"
    }}

    Hãy cố gắng điền đầy đủ thông tin contact nếu có thể tìm thấy trong nội dung hoặc từ URL website của công ty.
    """ 
    time.sleep(2)
    response = model.generate_content(prompt)
    #print(response.text)
    return await process_gemini_response(response.text)

async def main(): 
    browser_config = BrowserConfig()
    run_config = CrawlerRunConfig(
        only_text=True,
        exclude_external_links=True,
        remove_forms=True,
        excluded_tags=["script", "style"],
        cache_mode=CacheMode.BYPASS,
        keep_attrs=["src", "class"],
        css_selector=".col-md-9",  
    )
    
    all_jobs = []  # Tạo list để lưu tất cả job_info

    async with AsyncWebCrawler(config=browser_config) as crawler:
        
        # duyệt từng url trong danh sách
        for url in urls:
            try:    
                result = await crawler.arun(url=url, config=run_config)
                print(f"Crawl successful: {result.success}")
                print(f"Status code: {result.status_code}")
                job_info = await get_job_info(result.markdown, url)
                if job_info:
                    if isinstance(job_info, list):
                        all_jobs.extend(job_info)  # nếu job_info là list
                    else:
                        all_jobs.append(job_info)  # nếu job_info là dict đơn
            except Exception as e:
                print(f"Error processing URL {url}: {str(e)}")
    
    # Export Data to new excel file
    if all_jobs:
        export_to_excel(all_jobs)
    else:
        print("⚠️ Không có dữ liệu để export.")
        
    # Sau khi crawl xong -> ghi dữ liệu vào Excel
def export_to_excel(data):
    now = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_file = f"jobs_data_{now}.xlsx"

    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)

    # Format lại file Excel
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Tô màu dòng tiêu đề
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Auto adjust width cho từng cột
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = (length + 2)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    wb.save(output_file)
    print(f"✅ Export thành công file Excel: {output_file}")

if __name__ == "__main__":
    asyncio.run(main()) 
