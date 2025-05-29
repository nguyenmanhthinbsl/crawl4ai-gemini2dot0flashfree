import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from job_deduplicator import JobDeduplicator
from typing import List, Dict, Any

class JobExporter:
    """Handles exporting job data to Excel with duplicate highlighting."""
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]]):
        now = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_file = f"jobs_data_{now}.xlsx"
        df = pd.DataFrame(data)
        df.to_excel(output_file, index=False)
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        duplicate_colors = [
            "FFB8E0F7", "FFFFD3B6", "FFD5F5E3", "FFF2D7D5", "FFFDEBD0", "FFE8DAEF", "FFDAEEF3", "FFD4EFDF"
        ]
        duplicate_groups = JobDeduplicator.find_duplicate_jobs(data)
        color_mapping = {}
        for i, (key, indices) in enumerate(duplicate_groups.items()):
            color = duplicate_colors[i % len(duplicate_colors)]
            for idx in indices:
                color_mapping[idx] = color
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            if row_idx - 1 in color_mapping:
                fill = PatternFill(start_color=color_mapping[row_idx - 1], end_color=color_mapping[row_idx - 1], fill_type="solid")
                for cell in row:
                    cell.fill = fill
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = (length + 2)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        legend_row = ws.max_row + 2
        ws.cell(row=legend_row, column=1, value="Chú thích màu:").font = Font(bold=True)
        legend_entries = []
        for i, (key, indices) in enumerate(duplicate_groups.items()):
            company_title = key.split('_')
            if len(company_title) == 2:
                company, title = company_title
                legend_text = f"Công việc trùng lặp {i+1}: {title.title()} - {company.title()}"
                legend_entries.append((legend_text, duplicate_colors[i % len(duplicate_colors)]))
        for i, (text, color) in enumerate(legend_entries):
            row = legend_row + i + 1
            cell = ws.cell(row=row, column=1, value=text)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        wb.save(output_file)
        print(f"✅ Export thành công file Excel: {output_file}")
        if duplicate_groups:
            print(f"🔍 Đã tìm thấy {len(duplicate_groups)} nhóm việc làm trùng lặp") 